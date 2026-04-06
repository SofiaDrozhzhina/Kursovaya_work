"""
Система управления курсами — Flask + PostgreSQL
Запуск:
    pip install flask psycopg2-binary flask-login werkzeug
    python app.py
"""

from flask import Flask, render_template, redirect, url_for, request, flash, session, jsonify, send_file
from functools import wraps
import psycopg2
import psycopg2.extras
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime
import os
import io

app = Flask(__name__)

# ─── Upload helpers ──────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Картинки курсов хранятся в проекте (static/course_images), без БД.
COURSE_IMAGE_MAP = {
    "основы python": "course_images/python.svg",
    "базы данных и sql": "course_images/sql.svg",
    "веб-разработка на flask": "course_images/flask.svg",
    "математический анализ": "course_images/math_analysis.svg",
    "линейная алгебра": "course_images/linear_algebra.svg",
    "машинное обучение": "course_images/ml.svg",
    "экономика предприятия": "course_images/economy.svg",
    "управление проектами": "course_images/project_management.svg",
    "дискретная математика": "course_images/discrete_math.svg",
    "алгоритмы и структуры данных": "course_images/algorithms.svg",
}

def course_image_for(title: str | None) -> str:
    if not title:
        return "course_images/default.svg"
    return COURSE_IMAGE_MAP.get(title.strip().lower(), "course_images/default.svg")
ALLOWED_IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".webp"}

def _allowed_image(filename: str) -> bool:
    ext = os.path.splitext(filename.lower())[1]
    return ext in ALLOWED_IMAGE_EXTS

def save_course_image(file_storage, old_rel_path=None):
    """Сохраняет картинку курса в static/uploads/courses и возвращает относительный путь (для БД)."""
    if not file_storage or not getattr(file_storage, "filename", ""):
        return None
    filename = secure_filename(file_storage.filename)
    if not filename:
        return None
    if not _allowed_image(filename):
        raise ValueError("Разрешены только изображения: png, jpg, jpeg, webp.")
    # Ограничение ~5MB
    try:
        file_storage.stream.seek(0, os.SEEK_END)
        size = file_storage.stream.tell()
        file_storage.stream.seek(0)
        if size > 5 * 1024 * 1024:
            raise ValueError("Файл слишком большой (максимум 5MB).")
    except Exception:
        pass

    uploads_dir = os.path.join(BASE_DIR, "static", "uploads", "courses")
    os.makedirs(uploads_dir, exist_ok=True)

    ts = datetime.now().strftime("%Y%m%d%H%M%S")
    name, ext = os.path.splitext(filename)
    final_name = f"{name}_{ts}{ext}"
    abs_path = os.path.join(uploads_dir, final_name)
    file_storage.save(abs_path)

    if old_rel_path:
        try:
            old_abs = os.path.join(BASE_DIR, "static", old_rel_path.replace("/", os.sep))
            if os.path.exists(old_abs):
                os.remove(old_abs)
        except Exception:
            pass

    return f"uploads/courses/{final_name}"
app.secret_key = os.environ.get("SECRET_KEY", "super-secret-dev-key-change-in-prod")

# ── Jinja2 helpers ────────────────────────────────────────────────────────────

@app.context_processor
def inject_globals():
    return {"now": datetime.now(), "course_image_for": course_image_for}

app.jinja_env.globals["enumerate"] = enumerate

# ─── Database connection ─────────────────────────────────────────────────────

DB_CONFIG = {
    "host":     os.environ.get("DB_HOST",     "localhost"),
    "port":     int(os.environ.get("DB_PORT", 5432)),
    "dbname":   os.environ.get("DB_NAME",     "courses_db"),
    "user":     os.environ.get("DB_USER",     "postgres"),
    "password": os.environ.get("DB_PASSWORD", "1234"),
}

def get_db():
    conn = psycopg2.connect(**DB_CONFIG)
    conn.autocommit = False
    return conn

def query(sql, params=(), fetchone=False, fetchall=False, commit=False):
    conn = get_db()
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(sql, params)
        result = None
        if fetchone:
            result = cur.fetchone()
        elif fetchall:
            result = cur.fetchall()
        if commit:
            conn.commit()
        return result
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

# ─── Auth helpers ────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if session.get("role") not in roles:
                flash("Недостаточно прав доступа.", "error")
                return redirect(url_for("dashboard"))
            return f(*args, **kwargs)
        return decorated
    return decorator

def current_user():
    if "user_id" not in session:
        return None
    return {
        "id":        session["user_id"],
        "role":      session["role"],
        "full_name": session["full_name"],
        "username":  session["username"],
    }

# ─── Routes: Auth ────────────────────────────────────────────────────────────

@app.route("/")
def index():
    if "user_id" in session:
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = query("SELECT * FROM users WHERE username=%s", (username,), fetchone=True)
        if user and check_password_hash(user["password"], password):
            session["user_id"]   = user["id"]
            session["role"]      = user["role"]
            session["full_name"] = user["full_name"] or username
            session["username"]  = user["username"]
            return redirect(url_for("dashboard"))
        flash("Неверный логин или пароль.", "error")
    return render_template("login.html")

@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        username  = request.form.get("username", "").strip()
        full_name = request.form.get("full_name", "").strip()
        email     = request.form.get("email", "").strip()
        password  = request.form.get("password", "")
        password2 = request.form.get("password2", "")
        role      = request.form.get("role", "student")

        if not username or not password or not full_name:
            flash("Заполните все обязательные поля.", "error")
        elif len(username) < 3:
            flash("Логин должен содержать минимум 3 символа.", "error")
        elif password != password2:
            flash("Пароли не совпадают.", "error")
        elif len(password) < 6:
            flash("Пароль должен быть не менее 6 символов.", "error")
        elif email and "@" not in email:
            flash("Введите корректный email.", "error")
        elif role not in ("student", "teacher"):
            flash("Недопустимая роль.", "error")
        else:
            existing = query("SELECT id FROM users WHERE username=%s", (username,), fetchone=True)
            if existing:
                flash(f"Логин «{username}» уже занят.", "error")
            else:
                hashed = generate_password_hash(password)
                query(
                    "INSERT INTO users (username, password, role, full_name, email) VALUES (%s,%s,%s,%s,%s) RETURNING id",
                    (username, hashed, role, full_name, email),
                    commit=True
                )
                user = query("SELECT id FROM users WHERE username=%s", (username,), fetchone=True)
                uid = user["id"]
                if role == "teacher":
                    query("INSERT INTO teachers (user_id) VALUES (%s)", (uid,), commit=True)
                elif role == "student":
                    query("INSERT INTO students (user_id, enrollment_year) VALUES (%s,%s)",
                          (uid, datetime.now().year), commit=True)
                flash("Аккаунт создан! Войдите в систему.", "success")
                return redirect(url_for("login"))
    return render_template("register.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# ─── Dashboard ───────────────────────────────────────────────────────────────

@app.route("/dashboard")
@login_required
def dashboard():
    user = current_user()
    role = user["role"]
    stats = {}

    if role in ("admin", "teacher"):
        stats["students"]  = query("SELECT COUNT(*) as n FROM students", fetchone=True)["n"]
        stats["courses"]   = query("SELECT COUNT(*) as n FROM courses",  fetchone=True)["n"]
        stats["teachers"]  = query("SELECT COUNT(*) as n FROM teachers", fetchone=True)["n"]
        stats["avg_grade"] = query("SELECT ROUND(AVG(grade)::numeric,1) as n FROM grades", fetchone=True)["n"] or 0
        recent = query("""
            SELECT u.full_name AS student, co.title AS course,
                   e.enrolled_at, e.status
            FROM enrollments e
            JOIN students s ON e.student_id=s.id
            JOIN users u ON s.user_id=u.id
            JOIN courses co ON e.course_id=co.id
            ORDER BY e.enrolled_at DESC LIMIT 10
        """, fetchall=True)
    else:
        sid = query("SELECT id FROM students WHERE user_id=%s", (user["id"],), fetchone=True)
        if sid:
            sid = sid["id"]
            stats["enrolled"]  = query("SELECT COUNT(*) as n FROM enrollments WHERE student_id=%s", (sid,), fetchone=True)["n"]
            stats["completed"] = query("SELECT COUNT(*) as n FROM enrollments WHERE student_id=%s AND status='completed'", (sid,), fetchone=True)["n"]
            stats["avg_grade"] = query("""
                SELECT ROUND(AVG(g.grade)::numeric,1) as n FROM grades g
                JOIN enrollments e ON g.enrollment_id=e.id WHERE e.student_id=%s
            """, (sid,), fetchone=True)["n"] or 0
        recent = query("""
            SELECT co.title AS course, e.enrolled_at, e.status, g.grade
            FROM enrollments e
            JOIN courses co ON e.course_id=co.id
            LEFT JOIN grades g ON g.enrollment_id=e.id
            JOIN students s ON e.student_id=s.id
            WHERE s.user_id=%s ORDER BY e.enrolled_at DESC LIMIT 8
        """, (user["id"],), fetchall=True)

    return render_template("dashboard.html", user=user, stats=stats, recent=recent)

# ─── Users (admin) ──────────────────────────────────────────────────────────

@app.route("/users")
@login_required
@role_required("admin")
def users():
    rows = query("""
        SELECT id, username, full_name, email, role, created_at FROM users ORDER BY id
    """, fetchall=True)
    return render_template("users.html", user=current_user(), rows=rows)

@app.route("/users/delete/<int:uid>", methods=["POST"])
@login_required
@role_required("admin")
def delete_user(uid):
    if uid == session["user_id"]:
        flash("Нельзя удалить себя.", "error")
    else:
        query("DELETE FROM users WHERE id=%s", (uid,), commit=True)
        flash("Пользователь удалён.", "success")
    return redirect(url_for("users"))

@app.route("/users/add", methods=["POST"])
@login_required
@role_required("admin")
def add_user():
    username  = request.form.get("username", "").strip()
    full_name = request.form.get("full_name", "").strip()
    email     = request.form.get("email", "").strip()
    password  = request.form.get("password", "")
    role      = request.form.get("role", "student")

    # Валидация
    if not username or not password or not full_name:
        flash("Логин, имя и пароль обязательны.", "error")
        return redirect(url_for("users"))
    if len(username) < 3:
        flash("Логин должен содержать минимум 3 символа.", "error")
        return redirect(url_for("users"))
    if len(password) < 6:
        flash("Пароль должен содержать минимум 6 символов.", "error")
        return redirect(url_for("users"))
    if email and "@" not in email:
        flash("Введите корректный email.", "error")
        return redirect(url_for("users"))

    # Администратор должен быть только один
    if role == "admin":
        existing_admin = query("SELECT id FROM users WHERE role='admin'", fetchone=True)
        if existing_admin:
            flash("В системе уже есть администратор. Допускается только один администратор.", "error")
            return redirect(url_for("users"))

    existing = query("SELECT id FROM users WHERE username=%s", (username,), fetchone=True)
    if existing:
        flash(f"Логин «{username}» уже занят.", "error")
    else:
        query("INSERT INTO users (username,password,role,full_name,email) VALUES (%s,%s,%s,%s,%s)",
              (username, generate_password_hash(password), role, full_name, email), commit=True)
        uid = query("SELECT id FROM users WHERE username=%s", (username,), fetchone=True)["id"]
        if role == "teacher":
            query("INSERT INTO teachers (user_id) VALUES (%s)", (uid,), commit=True)
        elif role == "student":
            query("INSERT INTO students (user_id, enrollment_year) VALUES (%s,%s)",
                  (uid, datetime.now().year), commit=True)
        flash(f"Пользователь «{full_name}» успешно добавлен.", "success")
    return redirect(url_for("users"))


@app.route("/users/edit/<int:uid>", methods=["POST"])
@login_required
@role_required("admin")
def edit_user(uid):
    username  = request.form.get("username", "").strip()
    full_name = request.form.get("full_name", "").strip()
    email     = request.form.get("email", "").strip()

    if not username or not full_name:
        flash("Логин и имя обязательны.", "error")
        return redirect(url_for("users"))
    if len(username) < 3:
        flash("Логин должен содержать минимум 3 символа.", "error")
        return redirect(url_for("users"))
    if email and "@" not in email:
        flash("Введите корректный email.", "error")
        return redirect(url_for("users"))

    existing = query("SELECT id FROM users WHERE username=%s AND id<>%s", (username, uid), fetchone=True)
    if existing:
        flash(f"Логин «{username}» уже занят.", "error")
        return redirect(url_for("users"))

    query("UPDATE users SET username=%s, full_name=%s, email=%s WHERE id=%s",
          (username, full_name, email or None, uid), commit=True)
    flash("Пользователь обновлён.", "success")
    return redirect(url_for("users"))

# ─── Students ───────────────────────────────────────────────────────────────

@app.route("/students")
@login_required
@role_required("admin", "teacher")
def students():
    search = request.args.get("q", "")
    rows = query("""
        SELECT s.id, u.full_name, u.username, u.email,
               s.group_name, s.enrollment_year,
               COUNT(DISTINCT e.id) AS total_courses,
               COUNT(DISTINCT CASE WHEN e.status='completed' THEN e.id END) AS completed,
               ROUND(AVG(g.grade)::numeric,1) AS avg_grade
        FROM students s
        JOIN users u ON s.user_id=u.id
        LEFT JOIN enrollments e ON e.student_id=s.id
        LEFT JOIN grades g ON g.enrollment_id=e.id
        WHERE u.full_name ILIKE %s OR u.username ILIKE %s
        GROUP BY s.id, u.full_name, u.username, u.email, s.group_name, s.enrollment_year
        ORDER BY u.full_name
    """, (f"%{search}%", f"%{search}%"), fetchall=True)
    return render_template("students.html", user=current_user(), rows=rows, search=search)

@app.route("/students/<int:sid>/edit", methods=["POST"])
@login_required
@role_required("admin")
def edit_student(sid):
    full_name = request.form.get("full_name", "").strip()
    email     = request.form.get("email", "").strip()
    group     = request.form.get("group_name", "").strip()
    year_raw  = request.form.get("enrollment_year", "").strip()

    if not full_name:
        flash("Имя студента не может быть пустым.", "error")
        return redirect(url_for("students"))
    if email and "@" not in email:
        flash("Введите корректный email.", "error")
        return redirect(url_for("students"))
    year = None
    if year_raw:
        try:
            year = int(year_raw)
            if year < 2000 or year > datetime.now().year:
                raise ValueError
        except ValueError:
            flash(f"Год поступления должен быть числом от 2000 до {datetime.now().year}.", "error")
            return redirect(url_for("students"))

    query("UPDATE students SET group_name=%s, enrollment_year=%s WHERE id=%s",
          (group or None, year, sid), commit=True)
    query("UPDATE users SET full_name=%s, email=%s WHERE id=(SELECT user_id FROM students WHERE id=%s)",
          (full_name, email or None, sid), commit=True)
    flash("Данные студента успешно обновлены.", "success")
    return redirect(url_for("students"))

@app.route("/students/<int:sid>/delete", methods=["POST"])
@login_required
@role_required("admin")
def delete_student(sid):
    info = query("SELECT u.full_name, u.id AS uid FROM students s JOIN users u ON s.user_id=u.id WHERE s.id=%s",
                 (sid,), fetchone=True)
    if not info:
        flash("Студент не найден.", "error")
        return redirect(url_for("students"))
    query("DELETE FROM users WHERE id=%s", (info["uid"],), commit=True)
    flash(f"Студент «{info['full_name']}» удалён.", "success")
    return redirect(url_for("students"))

# ─── Teachers ───────────────────────────────────────────────────────────────

@app.route("/teachers")
@login_required
@role_required("admin", "teacher")
def teachers():
    rows = query("""
        SELECT t.id, u.full_name, u.username, u.email,
               t.department, t.degree,
               COUNT(DISTINCT c.id) AS courses_count,
               COUNT(DISTINCT e.student_id) AS students_count
        FROM teachers t
        JOIN users u ON t.user_id=u.id
        LEFT JOIN courses c ON c.teacher_id=t.id
        LEFT JOIN enrollments e ON e.course_id=c.id AND e.status='active'
        GROUP BY t.id, u.full_name, u.username, u.email, t.department, t.degree
        ORDER BY u.full_name
    """, fetchall=True)
    return render_template("teachers.html", user=current_user(), rows=rows)

@app.route("/teachers/<int:tid>/edit", methods=["POST"])
@login_required
@role_required("admin")
def edit_teacher(tid):
    full_name  = request.form.get("full_name", "").strip()
    email      = request.form.get("email", "").strip()
    department = request.form.get("department", "").strip()
    degree     = request.form.get("degree", "").strip()

    if not full_name:
        flash("ФИО преподавателя не может быть пустым.", "error")
        return redirect(url_for("teachers"))
    if email and "@" not in email:
        flash("Введите корректный email.", "error")
        return redirect(url_for("teachers"))

    query("UPDATE teachers SET department=%s, degree=%s WHERE id=%s",
          (department or None, degree or None, tid), commit=True)
    query("UPDATE users SET full_name=%s, email=%s WHERE id=(SELECT user_id FROM teachers WHERE id=%s)",
          (full_name, email or None, tid), commit=True)

    flash("Данные преподавателя обновлены.", "success")
    return redirect(url_for("teachers"))

@app.route("/teachers/<int:tid>/delete", methods=["POST"])
@login_required
@role_required("admin")
def delete_teacher(tid):
    info = query("SELECT u.full_name, u.id AS uid FROM teachers t JOIN users u ON t.user_id=u.id WHERE t.id=%s",
                 (tid,), fetchone=True)
    if not info:
        flash("Преподаватель не найден.", "error")
        return redirect(url_for("teachers"))
    # Снимаем привязку курсов перед удалением
    query("UPDATE courses SET teacher_id=NULL WHERE teacher_id=%s", (tid,), commit=True)
    query("DELETE FROM users WHERE id=%s", (info["uid"],), commit=True)
    flash(f"Преподаватель «{info['full_name']}» удалён.", "success")
    return redirect(url_for("teachers"))

# ─── Courses ─────────────────────────────────────────────────────────────────

@app.route("/courses")
@login_required
def courses():
    user = current_user()
    rows = query("""
        SELECT co.id, co.title, co.description, u.full_name AS teacher_name,
               co.max_students,
               COUNT(DISTINCT CASE WHEN e.status<>'dropped' THEN e.id END) AS enrolled,
               COUNT(DISTINCT CASE WHEN e.status='completed' THEN e.id END) AS completed,
               ROUND(AVG(g.grade)::numeric,1) AS avg_grade,
               co.created_at
        FROM courses co
        LEFT JOIN teachers t ON co.teacher_id=t.id
        LEFT JOIN users u ON t.user_id=u.id
        LEFT JOIN enrollments e ON e.course_id=co.id
        LEFT JOIN grades g ON g.enrollment_id=e.id
        GROUP BY co.id, co.title, co.description, u.full_name, co.max_students, co.created_at
        ORDER BY co.title
    """, fetchall=True)
    teacher_list = query("""
        SELECT t.id, u.full_name FROM teachers t JOIN users u ON t.user_id=u.id ORDER BY u.full_name
    """, fetchall=True)
    return render_template("courses.html", user=user, rows=rows, teacher_list=teacher_list)

@app.route("/courses/add", methods=["POST"])
@login_required
@role_required("admin")
def add_course():
    title = request.form.get("title", "").strip()
    if not title:
        flash("Название курса обязательно.", "error")
        return redirect(url_for("courses"))
    if len(title) < 3:
        flash("Название курса должно содержать минимум 3 символа.", "error")
        return redirect(url_for("courses"))

    max_s_raw = request.form.get("max_students", "30")
    try:
        max_s = int(max_s_raw)
        if max_s < 1 or max_s > 500:
            raise ValueError
    except ValueError:
        flash("Максимальное количество студентов должно быть числом от 1 до 500.", "error")
        return redirect(url_for("courses"))

    query("""
        INSERT INTO courses (title, description, teacher_id, max_students)
        VALUES (%s,%s,%s,%s)
    """, (
        title,
        request.form.get("description", "").strip() or None,
        request.form.get("teacher_id") or None,
        max_s
    ), commit=True)
    flash(f"Курс «{title}» успешно добавлен.", "success")
    return redirect(url_for("courses"))

@app.route("/courses/<int:cid>/edit", methods=["POST"])
@login_required
@role_required("admin")
def edit_course(cid):
    title = request.form.get("title", "").strip()
    if not title:
        flash("Название курса обязательно.", "error")
        return redirect(url_for("courses"))

    max_s_raw = request.form.get("max_students", "30")
    try:
        max_s = int(max_s_raw)
        if max_s < 1 or max_s > 500:
            raise ValueError
    except ValueError:
        flash("Максимальное количество студентов должно быть числом от 1 до 500.", "error")
        return redirect(url_for("courses"))

    query(
        "UPDATE courses SET title=%s, description=%s, teacher_id=%s, max_students=%s WHERE id=%s",
        (
            title,
            request.form.get("description", "").strip() or None,
            request.form.get("teacher_id") or None,
            max_s,
            cid,
        ),
        commit=True,
    )
    flash("Курс обновлён.", "success")
    return redirect(url_for("courses"))

@app.route("/courses/<int:cid>/delete", methods=["POST"])
@login_required
@role_required("admin")
def delete_course(cid):
    query("DELETE FROM courses WHERE id=%s", (cid,), commit=True)
    flash("Курс удалён.", "success")
    return redirect(url_for("courses"))

# ─── Grades ──────────────────────────────────────────────────────────────────

@app.route("/grades")
@login_required
@role_required("admin", "teacher")
def grades():
    user = current_user()
    if user["role"] == "teacher":
        tid = query("SELECT id FROM teachers WHERE user_id=%s", (user["id"],), fetchone=True)
        tid = tid["id"] if tid else -1
        rows = query("""
            SELECT e.id, u.full_name AS student, co.title AS course,
                   g.grade, g.graded_at, g.comment, e.status
            FROM enrollments e
            JOIN students s ON e.student_id=s.id
            JOIN users u ON s.user_id=u.id
            JOIN courses co ON e.course_id=co.id
            LEFT JOIN grades g ON g.enrollment_id=e.id
            WHERE co.teacher_id=%s
            ORDER BY u.full_name, co.title
        """, (tid,), fetchall=True)
    else:
        rows = query("""
            SELECT e.id, u.full_name AS student, co.title AS course,
                   g.grade, g.graded_at, g.comment, e.status
            FROM enrollments e
            JOIN students s ON e.student_id=s.id
            JOIN users u ON s.user_id=u.id
            JOIN courses co ON e.course_id=co.id
            LEFT JOIN grades g ON g.enrollment_id=e.id
            ORDER BY u.full_name, co.title
        """, fetchall=True)
    return render_template("grades.html", user=user, rows=rows)

@app.route("/grades/set", methods=["POST"])
@login_required
@role_required("admin", "teacher")
def set_grade():
    eid_raw   = request.form.get("enrollment_id", "")
    grade_raw = request.form.get("grade", "")
    comment   = request.form.get("comment", "").strip()

    if not eid_raw or not grade_raw:
        flash("Заполните все обязательные поля.", "error")
        return redirect(url_for("grades"))
    try:
        eid   = int(eid_raw)
        grade = float(grade_raw.replace(",", "."))
        if grade < 0 or grade > 100:
            raise ValueError
    except ValueError:
        flash("Оценка должна быть числом от 0 до 100.", "error")
        return redirect(url_for("grades"))

    enrollment = query("SELECT id, course_id FROM enrollments WHERE id=%s", (eid,), fetchone=True)
    if not enrollment:
        flash("Запись на курс не найдена.", "error")
        return redirect(url_for("grades"))

    user = current_user()
    if user["role"] == "teacher":
        tid = query("SELECT id FROM teachers WHERE user_id=%s", (user["id"],), fetchone=True)
        tid = tid["id"] if tid else -1
        allowed = query("SELECT id FROM courses WHERE id=%s AND teacher_id=%s",
                        (enrollment["course_id"], tid), fetchone=True)
        if not allowed:
            flash("Нельзя выставлять оценку по чужому курсу.", "error")
            return redirect(url_for("grades"))

    query("""
        INSERT INTO grades (enrollment_id, grade, comment, graded_at)
        VALUES (%s,%s,%s,NOW())
        ON CONFLICT (enrollment_id) DO UPDATE
        SET grade=EXCLUDED.grade, comment=EXCLUDED.comment, graded_at=NOW()
    """, (eid, grade, comment), commit=True)
    query("UPDATE enrollments SET status='completed' WHERE id=%s", (eid,), commit=True)
    flash(f"Оценка {grade} успешно сохранена.", "success")
    return redirect(url_for("grades"))

    # Проверяем что запись существует
    enrollment = query("SELECT id FROM enrollments WHERE id=%s", (eid,), fetchone=True)
    if not enrollment:
        flash("Запись на курс не найдена.", "error")
        return redirect(url_for("grades"))

    query("""
        INSERT INTO grades (enrollment_id, grade, comment, graded_at)
        VALUES (%s,%s,%s,NOW())
        ON CONFLICT (enrollment_id) DO UPDATE
        SET grade=EXCLUDED.grade, comment=EXCLUDED.comment, graded_at=NOW()
    """, (eid, grade, comment), commit=True)
    query("UPDATE enrollments SET status='completed' WHERE id=%s", (eid,), commit=True)
    flash(f"Оценка {grade} успешно сохранена.", "success")
    return redirect(url_for("grades"))


@app.route("/grades/delete/<int:eid>", methods=["POST"])
@login_required
@role_required("admin")
def delete_grade(eid):
    g = query("SELECT id FROM grades WHERE enrollment_id=%s", (eid,), fetchone=True)
    if not g:
        flash("Оценка не найдена.", "error")
        return redirect(url_for("grades"))

    query("DELETE FROM grades WHERE enrollment_id=%s", (eid,), commit=True)
    query("UPDATE enrollments SET status='active' WHERE id=%s", (eid,), commit=True)

    flash("Оценка удалена.", "success")
    return redirect(url_for("grades"))

# ─── Statistics ──────────────────────────────────────────────────────────────

@app.route("/statistics")
@login_required
@role_required("admin", "teacher")
def statistics():
    user = current_user()

    # Students
    st = {}
    st["total"]    = query("SELECT COUNT(*) AS n FROM students", fetchone=True)["n"]
    st["active"]   = query("""SELECT COUNT(DISTINCT s.id) AS n FROM students s
                              JOIN enrollments e ON e.student_id=s.id WHERE e.status='active'""", fetchone=True)["n"]
    st["avg"]      = query("SELECT ROUND(AVG(grade)::numeric,2) AS n FROM grades", fetchone=True)["n"] or 0
    st["max"]      = query("SELECT MAX(grade) AS n FROM grades", fetchone=True)["n"] or 0
    st["min"]      = query("SELECT MIN(grade) AS n FROM grades", fetchone=True)["n"] or 0
    st["no_grade"] = query("""SELECT COUNT(DISTINCT s.id) AS n FROM students s
                              JOIN enrollments e ON e.student_id=s.id
                              LEFT JOIN grades g ON g.enrollment_id=e.id
                              WHERE g.grade IS NULL""", fetchone=True)["n"]
    st["top5"]     = query("""
        SELECT u.full_name, ROUND(AVG(g.grade)::numeric,1) AS avg
        FROM students s JOIN users u ON s.user_id=u.id
        JOIN enrollments e ON e.student_id=s.id
        JOIN grades g ON g.enrollment_id=e.id
        GROUP BY s.id, u.full_name ORDER BY avg DESC LIMIT 5
    """, fetchall=True)

    # Courses
    co = {}
    co["total"]    = query("SELECT COUNT(*) AS n FROM courses", fetchone=True)["n"]
    co["with_t"]   = query("SELECT COUNT(*) AS n FROM courses WHERE teacher_id IS NOT NULL", fetchone=True)["n"]
    co["avg_enr"]  = query("SELECT ROUND(AVG(cnt)::numeric,1) AS n FROM (SELECT COUNT(*) cnt FROM enrollments GROUP BY course_id) t", fetchone=True)["n"] or 0
    co["popular"]  = query("""
        SELECT co.title, COUNT(e.id) AS cnt FROM courses co
        LEFT JOIN enrollments e ON e.course_id=co.id
        GROUP BY co.id ORDER BY cnt DESC LIMIT 1
    """, fetchone=True)
    co["top5"]     = query("""
        SELECT co.title, ROUND(AVG(g.grade)::numeric,1) AS avg
        FROM courses co JOIN enrollments e ON e.course_id=co.id
        JOIN grades g ON g.enrollment_id=e.id
        GROUP BY co.id ORDER BY avg DESC LIMIT 5
    """, fetchall=True)

    # Teachers
    tc = {}
    tc["total"]      = query("SELECT COUNT(*) AS n FROM teachers", fetchone=True)["n"]
    tc["active"]     = query("SELECT COUNT(DISTINCT teacher_id) AS n FROM courses WHERE teacher_id IS NOT NULL", fetchone=True)["n"]
    tc["avg_courses"]= query("SELECT ROUND(AVG(cnt)::numeric,1) AS n FROM (SELECT COUNT(*) cnt FROM courses GROUP BY teacher_id) t", fetchone=True)["n"] or 0
    tc["top5"]       = query("""
        SELECT u.full_name, COUNT(DISTINCT e.student_id) AS students
        FROM teachers t JOIN users u ON t.user_id=u.id
        JOIN courses c ON c.teacher_id=t.id
        JOIN enrollments e ON e.course_id=c.id
        GROUP BY t.id, u.full_name ORDER BY students DESC LIMIT 5
    """, fetchall=True)

    # Grades distribution
    dist = query("""
        SELECT
          COUNT(*) FILTER (WHERE grade>=90) AS excellent,
          COUNT(*) FILTER (WHERE grade>=75 AND grade<90) AS good,
          COUNT(*) FILTER (WHERE grade>=60 AND grade<75) AS satisf,
          COUNT(*) FILTER (WHERE grade<60) AS fail,
          COUNT(*) AS total
        FROM grades WHERE grade IS NOT NULL
    """, fetchone=True)

    return render_template("statistics.html", user=user, st=st, co=co, tc=tc, dist=dist)

# ─── Student: My courses ──────────────────────────────────────────────────────

@app.route("/my-courses")
@login_required
@role_required("student", "teacher")
def my_courses():
    user = current_user()
    if user["role"] == "teacher":
        tid = query("SELECT id FROM teachers WHERE user_id=%s", (user["id"],), fetchone=True)
        rows = []
        if tid:
            rows = query("""
                SELECT co.id, co.title, co.description, co.max_students,
                       COUNT(DISTINCT CASE WHEN e.status<>'dropped' THEN e.id END) AS enrolled,
                       ROUND(AVG(g.grade)::numeric,1) AS avg_grade
                FROM courses co
                LEFT JOIN enrollments e ON e.course_id=co.id
                LEFT JOIN grades g ON g.enrollment_id=e.id
                WHERE co.teacher_id=%s GROUP BY co.id, co.title, co.description, co.max_students ORDER BY co.title
            """, (tid["id"],), fetchall=True)
    else:
        rows = query("""
            SELECT co.id, co.title AS title, u.full_name AS teacher, e.status, g.grade, e.enrolled_at
            FROM enrollments e
            JOIN courses co ON e.course_id=co.id
            LEFT JOIN teachers t ON co.teacher_id=t.id
            LEFT JOIN users u ON t.user_id=u.id
            LEFT JOIN grades g ON g.enrollment_id=e.id
            JOIN students s ON e.student_id=s.id
            WHERE s.user_id=%s ORDER BY e.enrolled_at DESC
        """, (user["id"],), fetchall=True)
    return render_template("my_courses.html", user=user, rows=rows)

@app.route("/enroll", methods=["GET", "POST"])
@login_required
@role_required("student")
def enroll():
    user = current_user()
    sid  = query("SELECT id FROM students WHERE user_id=%s", (user["id"],), fetchone=True)
    if not sid:
        flash("Профиль студента не найден.", "error")
        return redirect(url_for("dashboard"))
    sid = sid["id"]

    if request.method == "POST":
        cid = int(request.form.get("course_id"))

        # Проверка: существует ли курс
        course = query("SELECT id, title, max_students FROM courses WHERE id=%s", (cid,), fetchone=True)
        if not course:
            flash("Курс не найден.", "error")
            return redirect(url_for("enroll"))

        # Проверка: уже записан
        existing = query("SELECT id FROM enrollments WHERE student_id=%s AND course_id=%s",
                         (sid, cid), fetchone=True)
        if existing:
            flash("Вы уже записаны на этот курс.", "error")
            return redirect(url_for("enroll"))

        # Проверка: не превышен ли лимит мест
        enrolled_count = query(
            "SELECT COUNT(*) AS n FROM enrollments WHERE course_id=%s AND status<>\'dropped\'", (cid,), fetchone=True
        )["n"]
        if enrolled_count >= course["max_students"]:
            flash(f"К сожалению, курс «{course['title']}» уже набрал максимальное количество участников ({course['max_students']}).", "error")
            return redirect(url_for("enroll"))

        query("INSERT INTO enrollments (student_id, course_id) VALUES (%s,%s)",
              (sid, cid), commit=True)
        flash(f"Вы успешно записались на курс «{course['title']}»!", "success")
        return redirect(url_for("enroll"))

    enrolled_ids = {r["course_id"] for r in query(
        "SELECT course_id FROM enrollments WHERE student_id=%s", (sid,), fetchall=True)}
    courses_list = query("""
        SELECT co.id, co.title, co.description, u.full_name AS teacher,
               COUNT(DISTINCT CASE WHEN e.status<>'dropped' THEN e.id END) AS enrolled,
               co.max_students
        FROM courses co
        LEFT JOIN teachers t ON co.teacher_id=t.id
        LEFT JOIN users u ON t.user_id=u.id
        LEFT JOIN enrollments e ON e.course_id=co.id
        GROUP BY co.id, co.title, co.description, u.full_name, co.max_students
        ORDER BY co.title
    """, fetchall=True)
    return render_template("enroll.html", user=user, courses=courses_list, enrolled_ids=enrolled_ids)

@app.route("/my-grades")
@login_required
@role_required("student")
def my_grades():
    user = current_user()
    rows = query("""
        SELECT co.title AS course, u.full_name AS teacher,
               g.grade, g.graded_at, g.comment, e.status
        FROM enrollments e
        JOIN courses co ON e.course_id=co.id
        LEFT JOIN teachers t ON co.teacher_id=t.id
        LEFT JOIN users u ON t.user_id=u.id
        LEFT JOIN grades g ON g.enrollment_id=e.id
        JOIN students s ON e.student_id=s.id
        WHERE s.user_id=%s ORDER BY e.enrolled_at DESC
    """, (user["id"],), fetchall=True)
    avg = query("""
        SELECT ROUND(AVG(g.grade)::numeric,2) AS n FROM grades g
        JOIN enrollments e ON g.enrollment_id=e.id
        JOIN students s ON e.student_id=s.id
        WHERE s.user_id=%s
    """, (user["id"],), fetchone=True)["n"]
    return render_template("my_grades.html", user=user, rows=rows, avg=avg)

# ─── API: chart data ─────────────────────────────────────────────────────────

@app.route("/api/chart/grade-dist")
@login_required
def api_grade_dist():
    dist = query("""
        SELECT
          COUNT(*) FILTER (WHERE grade>=90) AS excellent,
          COUNT(*) FILTER (WHERE grade>=75 AND grade<90) AS good,
          COUNT(*) FILTER (WHERE grade>=60 AND grade<75) AS satisf,
          COUNT(*) FILTER (WHERE grade<60) AS fail
        FROM grades WHERE grade IS NOT NULL
    """, fetchone=True)
    return jsonify(dict(dist))

@app.route("/api/chart/enrollments-by-course")
@login_required
def api_enrollments_by_course():
    rows = query("""
        SELECT co.title, COUNT(e.id) AS cnt
        FROM courses co LEFT JOIN enrollments e ON e.course_id=co.id
        GROUP BY co.id ORDER BY cnt DESC LIMIT 8
    """, fetchall=True)
    return jsonify([dict(r) for r in rows])

# ─── Excel: Журнал успеваемости (преподаватель и админ) ──────────────────────

@app.route("/gradebook/excel")
@login_required
@role_required("teacher", "admin")
def download_gradebook_excel():
    """Журнал успеваемости в формате Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash("Установите openpyxl: pip install openpyxl", "error")
        return redirect(url_for("grades"))

    user = current_user()
    now  = datetime.now()

    status_map = {"completed": "Завершён", "active": "Активен", "dropped": "Отчислен"}

    is_admin = user["role"] == "admin"
    if is_admin:
        headers = ["№", "ФИО студента", "Группа", "Курс", "Преподаватель",
                   "Дата записи", "Статус", "Оценка", "Комментарий", "Дата оценки"]
        col_widths = [5, 30, 12, 35, 28, 14, 14, 10, 35, 14]
    else:
        headers = ["№", "ФИО студента", "Группа", "Курс",
                   "Дата записи", "Статус", "Оценка", "Комментарий", "Дата оценки"]
        col_widths = [5, 30, 12, 35, 14, 14, 10, 35, 14]

    max_cols = len(headers)
    last_col_letter = get_column_letter(max_cols)

    # Получаем данные
    if user["role"] == "teacher":
        tid = query("SELECT id FROM teachers WHERE user_id=%s", (user["id"],), fetchone=True)
        if not tid:
            flash("Профиль преподавателя не найден.", "error")
            return redirect(url_for("grades"))
        rows = query("""
            SELECT u.full_name AS student_name, s.group_name,
                   co.title AS course_title, e.enrolled_at,
                   e.status, g.grade, g.comment, g.graded_at
            FROM enrollments e
            JOIN students s ON e.student_id=s.id
            JOIN users u ON s.user_id=u.id
            JOIN courses co ON e.course_id=co.id
            LEFT JOIN grades g ON g.enrollment_id=e.id
            WHERE co.teacher_id=%s
            ORDER BY co.title, u.full_name
        """, (tid["id"],), fetchall=True)
        teacher_name = user["full_name"]
    else:
        rows = query("""
            SELECT u.full_name AS student_name, s.group_name,
                   co.title AS course_title, e.enrolled_at,
                   e.status, g.grade, g.comment, g.graded_at,
                   ut.full_name AS teacher_name
            FROM enrollments e
            JOIN students s ON e.student_id=s.id
            JOIN users u ON s.user_id=u.id
            JOIN courses co ON e.course_id=co.id
            LEFT JOIN grades g ON g.enrollment_id=e.id
            LEFT JOIN teachers t ON co.teacher_id=t.id
            LEFT JOIN users ut ON t.user_id=ut.id
            ORDER BY co.title, u.full_name
        """, fetchall=True)
        teacher_name = "Администратор"

    wb = Workbook()
    ws = wb.active
    ws.title = "Журнал успеваемости"

    # Стили
    PURPLE       = "4B3F8C"
    PURPLE_LIGHT = "E8E4F8"
    GRAY_LIGHT   = "F5F5F5"
    WHITE        = "FFFFFF"

    hdr_font  = Font(name="Arial", bold=True, color=WHITE, size=11)
    hdr_fill  = PatternFill("solid", fgColor=PURPLE)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sub_font  = Font(name="Arial", bold=True, size=10)
    sub_fill  = PatternFill("solid", fgColor=PURPLE_LIGHT)
    sub_align = Alignment(horizontal="center", vertical="center")

    body_font    = Font(name="Arial", size=10)
    body_align   = Alignment(vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Заголовок документа ──────────────────────────────────
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = "ЖУРНАЛ УСПЕВАЕМОСТИ СТУДЕНТОВ"
    ws["A1"].font  = Font(name="Arial", bold=True, size=14, color=PURPLE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{last_col_letter}2")
    ws["A2"] = f"Сформирован: {now.strftime('%d.%m.%Y %H:%M')}  |  {teacher_name}"
    ws["A2"].font      = Font(name="Arial", italic=True, size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    ws.append([])

    # ── Заголовки ────────────────────────────────────────────
    ws.append(headers)
    hdr_row = ws.max_row
    ws.row_dimensions[hdr_row].height = 36
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=hdr_row, column=col)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = border
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── Данные ───────────────────────────────────────────────
    prev_course = None
    for idx, r in enumerate(rows, 1):
        course = r["course_title"]
        if course != prev_course:
            ws.append([])
            sep_row = ws.max_row
            ws.merge_cells(start_row=sep_row, start_column=1, end_row=sep_row, end_column=max_cols)
            c = ws.cell(row=sep_row, column=1)
            c.value     = f"  📖 {course}"
            c.font      = sub_font
            c.fill      = sub_fill
            c.alignment = sub_align
            c.border    = border
            ws.row_dimensions[sep_row].height = 22
            prev_course = course

        grade = r.get("grade")
        grade_val = float(grade) if grade is not None else None

        if is_admin:
            data_row = [
                idx,
                r["student_name"],
                r["group_name"] or "",
                r["course_title"],
                r.get("teacher_name") or "",
                r["enrolled_at"].strftime("%d.%m.%Y") if r["enrolled_at"] else "",
                status_map.get(r["status"], r["status"]),
                grade_val,
                r.get("comment") or "",
                r["graded_at"].strftime("%d.%m.%Y") if r.get("graded_at") else "",
            ]
            grade_col = 8
            center_cols = {1, 6, 7, 8, 10}
        else:
            data_row = [
                idx,
                r["student_name"],
                r["group_name"] or "",
                r["course_title"],
                r["enrolled_at"].strftime("%d.%m.%Y") if r["enrolled_at"] else "",
                status_map.get(r["status"], r["status"]),
                grade_val,
                r.get("comment") or "",
                r["graded_at"].strftime("%d.%m.%Y") if r.get("graded_at") else "",
            ]
            grade_col = 7
            center_cols = {1, 5, 6, 7, 9}

        ws.append(data_row)
        row_i = ws.max_row
        ws.row_dimensions[row_i].height = 20

        for col_i, val in enumerate(data_row, 1):
            cell = ws.cell(row=row_i, column=col_i)
            cell.font   = body_font
            cell.border = border
            cell.fill   = PatternFill("solid", fgColor=WHITE if idx % 2 == 0 else GRAY_LIGHT)
            cell.alignment = center_align if col_i in center_cols else body_align

            if col_i == grade_col and grade_val is not None:
                if grade_val >= 90:
                    cell.font = Font(name="Arial", size=10, bold=True, color="166534")
                elif grade_val >= 75:
                    cell.font = Font(name="Arial", size=10, bold=True, color="0284C7")
                elif grade_val >= 60:
                    cell.font = Font(name="Arial", size=10, bold=True, color="D97706")
                else:
                    cell.font = Font(name="Arial", size=10, bold=True, color="DC2626")

    # ── Итоги ────────────────────────────────────────────────
    ws.append([])
    grades_only = [float(r["grade"]) for r in rows if r.get("grade") is not None]
    ws.append([
        "ИТОГО:",
        f"Записей: {len(rows)}",
        f"С оценкой: {len(grades_only)}",
        f"Средний балл: {round(sum(grades_only)/len(grades_only),1) if grades_only else '—'}",
        f"Отлично (90-100): {sum(1 for g in grades_only if g>=90)}",
        f"Хорошо (75-89): {sum(1 for g in grades_only if 75<=g<90)}",
        f"Удовл. (60-74): {sum(1 for g in grades_only if 60<=g<75)}",
        f"Неудовл. (<60): {sum(1 for g in grades_only if g<60)}",
    ])
    while ws.max_column < max_cols:
        ws.cell(row=ws.max_row, column=ws.max_column + 1, value="")

    total_r = ws.max_row
    ws.row_dimensions[total_r].height = 24
    for col in range(1, max_cols + 1):
        c = ws.cell(row=total_r, column=col)
        c.font  = Font(name="Arial", bold=True, size=10, color=PURPLE)
        c.fill  = PatternFill("solid", fgColor=PURPLE_LIGHT)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    ws.freeze_panes = f"A{hdr_row + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"Журнал_успеваемости_{now.strftime('%d%m%Y')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ─── Documents ───────────────────────────────────────────────────────────────

def _import_docs():
    """Ленивый импорт генератора документов"""
    try:
        from docs_generator import generate_consent, generate_study_certificate, generate_certificate
        return generate_consent, generate_study_certificate, generate_certificate
    except ImportError:
        return None, None, None

@app.route("/documents")
@login_required
@role_required("student")
def documents():
    user = current_user()
    sid  = query("SELECT id, group_name, enrollment_year FROM students WHERE user_id=%s",
                 (user["id"],), fetchone=True)
    completed_courses = []
    if sid:
        completed_courses = query("""
            SELECT e.id AS enrollment_id, co.title, e.status, g.grade, g.graded_at,
                   u.full_name AS teacher_name
            FROM enrollments e
            JOIN courses co ON e.course_id=co.id
            LEFT JOIN teachers t ON co.teacher_id=t.id
            LEFT JOIN users u ON t.user_id=u.id
            LEFT JOIN grades g ON g.enrollment_id=e.id
            JOIN students s ON e.student_id=s.id
            WHERE s.user_id=%s AND e.status='completed'
            ORDER BY g.graded_at DESC
        """, (user["id"],), fetchall=True)
    return render_template("documents.html", user=user,
                           student=sid, completed_courses=completed_courses)


@app.route("/documents/consent")
@login_required
@role_required("student")
def download_consent():
    user  = current_user()
    sinfo = query("SELECT group_name FROM students WHERE user_id=%s",
                  (user["id"],), fetchone=True)
    group = sinfo["group_name"] if sinfo else ""

    generate_consent, _, _ = _import_docs()
    if not generate_consent:
        flash("Модуль генерации документов не установлен. Выполните: pip install python-docx", "error")
        return redirect(url_for("documents"))

    doc_bytes = generate_consent(user["full_name"], "", group)
    return send_file(
        io.BytesIO(doc_bytes),
        as_attachment=True,
        download_name=f"Согласие_на_обработку_данных_{user['full_name'].replace(' ','_')}.docx",
        mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )


@app.route("/documents/study-certificate")
@login_required
@role_required("student")
def download_study_certificate():
    user  = current_user()
    sinfo = query("SELECT group_name, enrollment_year FROM students WHERE user_id=%s",
                  (user["id"],), fetchone=True)
    if not sinfo:
        flash("Профиль студента не найден.", "error")
        return redirect(url_for("documents"))

    courses = query("""
        SELECT co.title, e.status, g.grade
        FROM enrollments e
        JOIN courses co ON e.course_id=co.id
        LEFT JOIN grades g ON g.enrollment_id=e.id
        JOIN students s ON e.student_id=s.id
        WHERE s.user_id=%s
        ORDER BY co.title
    """, (user["id"],), fetchall=True)

    _, generate_study_certificate, _ = _import_docs()
    if not generate_study_certificate:
        flash("Модуль генерации документов не установлен. Выполните: pip install python-docx", "error")
        return redirect(url_for("documents"))

    doc_bytes = generate_study_certificate(
        student_name    = user["full_name"],
        group_name      = sinfo["group_name"] or "",
        enrollment_year = sinfo["enrollment_year"] or datetime.now().year,
        courses         = [dict(c) for c in courses]
    )
    return send_file(
        io.BytesIO(doc_bytes),
        as_attachment=True,
        download_name=f"Справка_об_обучении_{user['full_name'].replace(' ','_')}.docx",
        mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )


@app.route("/documents/certificate/<int:enrollment_id>")
@login_required
@role_required("student")
def download_certificate(enrollment_id):
    user = current_user()
    info = query("""
        SELECT co.title AS course_title, u.full_name AS teacher_name,
               g.grade, g.graded_at, e.status
        FROM enrollments e
        JOIN courses co ON e.course_id=co.id
        LEFT JOIN teachers t ON co.teacher_id=t.id
        LEFT JOIN users u ON t.user_id=u.id
        LEFT JOIN grades g ON g.enrollment_id=e.id
        JOIN students s ON e.student_id=s.id
        WHERE e.id=%s AND s.user_id=%s
    """, (enrollment_id, user["id"]), fetchone=True)

    if not info:
        flash("Запись не найдена.", "error")
        return redirect(url_for("documents"))
    if info["status"] != "completed":
        flash("Сертификат доступен только для завершённых курсов.", "error")
        return redirect(url_for("documents"))
    if info["grade"] is None:
        flash("Сертификат недоступен: оценка ещё не выставлена.", "error")
        return redirect(url_for("documents"))

    pdf_bytes = _generate_certificate_pdf(
        student_name  = user["full_name"],
        course_title  = info["course_title"],
        teacher_name  = info["teacher_name"] or "",
        grade         = float(info["grade"]),
        graded_at     = info["graded_at"]
    )
    safe_course = info["course_title"].replace(' ','_').replace('/','_')[:40]
    return send_file(
        io.BytesIO(pdf_bytes),
        as_attachment=True,
        download_name=f"Сертификат_{safe_course}_{user['full_name'].replace(' ','_')}.pdf",
        mimetype="application/pdf"
    )


def _generate_certificate_pdf(student_name, course_title, teacher_name,
                              grade, graded_at=None):
    """Генерация PDF сертификата через reportlab (кириллица + аккуратные звёзды + подпись директора)."""
    from math import cos, sin, pi
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, HRFlowable, Table,
        TableStyle, Image, Flowable
    )
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfbase.pdfmetrics import registerFontFamily
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    class StarRating(Flowable):
        def __init__(self, filled, total=5, size=8*mm, gap=2.2*mm,
                     fill_color=colors.HexColor("#C9A84C"),
                     empty_color=colors.HexColor("#E6D9A7")):
            super().__init__()
            self.filled = max(0, min(int(filled), int(total)))
            self.total = int(total)
            self.size = float(size)
            self.gap = float(gap)
            self.fill_color = fill_color
            self.empty_color = empty_color
            self.width = self.total * self.size + (self.total - 1) * self.gap
            self.height = self.size
            self.hAlign = "CENTER"

        def _star_points(self, cx, cy, outer_r, inner_r):
            pts = []
            angle = -pi / 2
            step = pi / 5
            for i in range(10):
                r = outer_r if i % 2 == 0 else inner_r
                x = cx + cos(angle) * r
                y = cy + sin(angle) * r
                pts.extend([x, y])
                angle += step
            return pts

        def draw(self):
            c = self.canv
            outer = self.size / 2
            inner = outer * 0.48
            baseline_y = self.size / 2

            for i in range(self.total):
                cx = outer + i * (self.size + self.gap)
                pts = self._star_points(cx, baseline_y, outer, inner)

                path = c.beginPath()
                path.moveTo(pts[0], pts[1])
                for j in range(2, len(pts), 2):
                    path.lineTo(pts[j], pts[j + 1])
                path.close()

                if i < self.filled:
                    c.setFillColor(self.fill_color)
                    c.setStrokeColor(self.fill_color)
                    c.drawPath(path, fill=1, stroke=1)
                else:
                    c.setFillColor(colors.white)
                    c.setStrokeColor(self.empty_color)
                    c.setLineWidth(1)
                    c.drawPath(path, fill=1, stroke=1)

    buf = io.BytesIO()
    page_w, page_h = landscape(A4)

    base_dir = os.path.dirname(os.path.abspath(__file__))
    fonts_dir = os.path.join(base_dir, "fonts")

    def pick_font(*names):
        candidates = []
        for name in names:
            candidates.append(os.path.join(fonts_dir, name))
            win_dir = os.environ.get("WINDIR")
            if win_dir:
                candidates.append(os.path.join(win_dir, "Fonts", name))
        for path in candidates:
            if path and os.path.exists(path):
                return path
        return None

    regular = pick_font("DejaVuSans.ttf") or pick_font("arial.ttf")
    bold = pick_font("DejaVuSans-Bold.ttf") or pick_font("arialbd.ttf") or regular
    italic = pick_font("DejaVuSans-Oblique.ttf") or pick_font("ariali.ttf") or regular
    bold_italic = pick_font("DejaVuSans-BoldOblique.ttf") or pick_font("arialbi.ttf") or bold

    if not regular:
        raise FileNotFoundError(
            "Не найден шрифт для PDF. Положите DejaVuSans.ttf в папку fonts/ "
            "или убедитесь, что на компьютере доступен Arial (C:\\Windows\\Fonts)."
        )

    pdfmetrics.registerFont(TTFont("AppSans", regular))
    pdfmetrics.registerFont(TTFont("AppSans-Bold", bold))
    pdfmetrics.registerFont(TTFont("AppSans-Italic", italic))
    pdfmetrics.registerFont(TTFont("AppSans-BoldItalic", bold_italic))
    registerFontFamily(
        "AppSans",
        normal="AppSans",
        bold="AppSans-Bold",
        italic="AppSans-Italic",
        boldItalic="AppSans-BoldItalic",
    )

    PURPLE       = colors.HexColor("#4B3F8C")
    PURPLE_LIGHT = colors.HexColor("#7B6DC5")
    GOLD         = colors.HexColor("#C9A84C")
    GOLD_LIGHT   = colors.HexColor("#E6D9A7")
    GREEN        = colors.HexColor("#16A34A")
    BLUE         = colors.HexColor("#0284C7")
    ORANGE       = colors.HexColor("#D97706")
    GRAY         = colors.HexColor("#64748B")
    LIGHT_BG     = colors.HexColor("#F5F3FF")

    if grade >= 90:
        level_text = "с отличием"
        level_color = GREEN
        stars = 5
    elif grade >= 75:
        level_text = "с хорошим результатом"
        level_color = BLUE
        stars = 4
    elif grade >= 60:
        level_text = "с удовлетворительным результатом"
        level_color = ORANGE
        stars = 3
    else:
        level_text = ""
        level_color = GRAY
        stars = 2

    months_ru = {1:'января',2:'февраля',3:'марта',4:'апреля',5:'мая',6:'июня',
                 7:'июля',8:'августа',9:'сентября',10:'октября',11:'ноября',12:'декабря'}
    if graded_at:
        date_str = f"{graded_at.day} {months_ru[graded_at.month]} {graded_at.year} г."
    else:
        n = datetime.now()
        date_str = f"{n.day} {months_ru[n.month]} {n.year} г."

    def s(name, **kw):
        return ParagraphStyle(name, **kw)

    sOrg    = s("org",    fontName="AppSans-Bold",   fontSize=11, textColor=PURPLE,       alignment=TA_CENTER, spaceAfter=2)
    sTitle  = s("title",  fontName="AppSans-Bold",   fontSize=36, textColor=PURPLE,       alignment=TA_CENTER, spaceAfter=4, leading=42)
    sSub    = s("sub",    fontName="AppSans-Italic", fontSize=14, textColor=PURPLE_LIGHT, alignment=TA_CENTER, spaceAfter=16)
    sLabel  = s("label",  fontName="AppSans",        fontSize=12, textColor=GRAY,         alignment=TA_CENTER, spaceAfter=4)
    sName   = s("name",   fontName="AppSans-Bold",   fontSize=28, textColor=colors.black, alignment=TA_CENTER, spaceAfter=4, leading=34)
    sCourse = s("course", fontName="AppSans-Bold",   fontSize=20, textColor=PURPLE,       alignment=TA_CENTER, spaceAfter=8, leading=26)
    sLevel  = s("level",  fontName="AppSans-Italic", fontSize=14, textColor=level_color,  alignment=TA_CENTER, spaceAfter=2)
    sGrade  = s("grade",  fontName="AppSans-Bold",   fontSize=18, textColor=level_color,  alignment=TA_CENTER, spaceAfter=3)
    sSigL   = s("sigl",   fontName="AppSans",        fontSize=10, textColor=colors.black, alignment=TA_LEFT, leading=12)
    sSigR   = s("sigr",   fontName="AppSans",        fontSize=10, textColor=colors.black, alignment=TA_RIGHT, leading=12)
    sMp     = s("mp",     fontName="AppSans-Bold",   fontSize=14, textColor=PURPLE,       alignment=TA_CENTER)

    margin = 2*cm
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=margin, rightMargin=margin,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    content_w = page_w - 2*margin

    def draw_border(canvas_obj, doc_obj):
        c2 = canvas_obj
        w, h = landscape(A4)
        c2.setStrokeColor(PURPLE)
        c2.setLineWidth(4)
        c2.rect(1*cm, 1*cm, w - 2*cm, h - 2*cm)
        c2.setStrokeColor(GOLD)
        c2.setLineWidth(1.5)
        c2.rect(1.3*cm, 1.3*cm, w - 2.6*cm, h - 2.6*cm)
        c2.setFillColor(PURPLE)
        for cx, cy in [(1*cm, 1*cm),(w-1*cm,1*cm),(1*cm,h-1*cm),(w-1*cm,h-1*cm)]:
            c2.circle(cx, cy, 3*mm, fill=1, stroke=0)
        c2.setFillColor(LIGHT_BG)
        c2.setFillAlpha(0.4)
        c2.rect(1.3*cm, 1.3*cm, w-2.6*cm, h-2.6*cm, fill=1, stroke=0)
        c2.setFillAlpha(1.0)

    story = []
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph("ГОСУДАРСТВЕННОЕ БЮДЖЕТНОЕ ОБРАЗОВАТЕЛЬНОЕ УЧРЕЖДЕНИЕ «УНИВЕРСИТЕТ»", sOrg))
    story.append(HRFlowable(width="80%", thickness=1.5, color=GOLD, spaceAfter=10))
    story.append(Paragraph("СЕРТИФИКАТ", sTitle))
    story.append(Paragraph("об успешном прохождении курса", sSub))
    story.append(HRFlowable(width="60%", thickness=1, color=PURPLE_LIGHT, spaceAfter=12))

    story.append(Paragraph("Настоящим подтверждается, что", sLabel))
    story.append(Paragraph(student_name, sName))
    story.append(Paragraph("успешно прошёл(-а) обучение по курсу", sLabel))
    story.append(Paragraph(f"«{course_title}»", sCourse))

    if level_text:
        story.append(Paragraph(level_text, sLevel))
    story.append(Paragraph(f"Итоговая оценка: {grade} баллов", sGrade))
    story.append(Spacer(1, 1.5*mm))
    story.append(StarRating(stars, total=5, size=7.5*mm, gap=2.5*mm, fill_color=GOLD, empty_color=GOLD_LIGHT))
    story.append(Spacer(1, 8*mm))

    story.append(HRFlowable(width="80%", thickness=1, color=GOLD, spaceBefore=0, spaceAfter=10))

    director_name = os.environ.get("DIRECTOR_NAME", "Иванов И.И.")
    sig_rel = os.environ.get("DIRECTOR_SIGNATURE", "static/signatures/director.png")
    sig_abs = os.path.join(base_dir, sig_rel.replace("/", os.sep))

    sign_block = []
    if os.path.exists(sig_abs):
        try:
            sig_img = Image(sig_abs, width=4.5*cm, height=1.9*cm)
            sign_block.append(sig_img)
            sign_block.append(Spacer(1, 1.2*mm))
        except Exception:
            pass

    sign_block.append(Paragraph("________________", sSigR))

    right_inner = Table([
        [Paragraph(f"Ректор:<br/><b>{director_name}</b>", sSigR)],
        [Spacer(1, 1.5*mm)],
        [sign_block],
    ], colWidths=[content_w*0.35])
    right_inner.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("ALIGN", (0,0), (-1,-1), "RIGHT"),
        ("LEFTPADDING",  (0,0), (-1,-1), 0),
        ("RIGHTPADDING", (0,0), (-1,-1), 0),
        ("TOPPADDING",   (0,0), (-1,-1), 0),
        ("BOTTOMPADDING",(0,0), (-1,-1), 0),
    ]))

    sig_data = [[
        Paragraph(f"Дата выдачи:<br/><b>{date_str}</b>", sSigL),
        Paragraph("М.П.", sMp),
        right_inner,
    ]]
    sig_table = Table(sig_data, colWidths=[content_w*0.35, content_w*0.3, content_w*0.35])
    sig_table.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "BOTTOM"),
        ("LEFTPADDING",  (0,0), (-1,-1), 0),
        ("RIGHTPADDING", (0,0), (-1,-1), 0),
        ("TOPPADDING",   (0,0), (-1,-1), 0),
        ("BOTTOMPADDING",(0,0), (-1,-1), 0),
    ]))
    story.append(sig_table)

    doc.build(story, onFirstPage=draw_border, onLaterPages=draw_border)
    buf.seek(0)
    return buf.read()




# ─── Документы для администратора ────────────────────────────────────────────

@app.route("/admin/documents")
@login_required
@role_required("admin")
def admin_documents():
    students_count  = query("SELECT COUNT(*) AS n FROM students",  fetchone=True)["n"]
    teachers_count  = query("SELECT COUNT(*) AS n FROM teachers",  fetchone=True)["n"]
    courses_count   = query("SELECT COUNT(*) AS n FROM courses",   fetchone=True)["n"]
    grades_count    = query("SELECT COUNT(*) AS n FROM grades WHERE grade IS NOT NULL", fetchone=True)["n"]
    students_list = query("""
        SELECT s.id, u.full_name, s.group_name, s.enrollment_year,
               COUNT(DISTINCT e.id) AS courses,
               ROUND(AVG(g.grade)::numeric,1) AS avg_grade
        FROM students s JOIN users u ON s.user_id=u.id
        LEFT JOIN enrollments e ON e.student_id=s.id
        LEFT JOIN grades g ON g.enrollment_id=e.id
        GROUP BY s.id, u.full_name, s.group_name, s.enrollment_year
        ORDER BY u.full_name
    """, fetchall=True)
    return render_template("admin_documents.html",
                           user=current_user(),
                           stats={"students":students_count,"teachers":teachers_count,
                                  "courses":courses_count,"grades":grades_count},
                           students_list=students_list)


@app.route("/admin/documents/consent/<int:student_id>")
@login_required
@role_required("admin")
def admin_download_consent(student_id):
    info = query("""SELECT u.full_name, u.email, s.group_name
                    FROM students s JOIN users u ON s.user_id=u.id WHERE s.id=%s""",
                 (student_id,), fetchone=True)
    if not info:
        flash("Студент не найден.", "error")
        return redirect(url_for("admin_documents"))
    generate_consent, _, _ = _import_docs()
    if not generate_consent:
        flash("Установите python-docx: pip install python-docx", "error")
        return redirect(url_for("admin_documents"))
    doc_bytes = generate_consent(info["full_name"], info["email"] or "", info["group_name"] or "")
    return send_file(io.BytesIO(doc_bytes), as_attachment=True,
                     download_name=f"Согласие_{info['full_name'].replace(' ','_')}.docx",
                     mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document")


@app.route("/admin/documents/study/<int:student_id>")
@login_required
@role_required("admin")
def admin_download_study(student_id):
    info = query("""SELECT u.full_name, s.group_name, s.enrollment_year
                    FROM students s JOIN users u ON s.user_id=u.id WHERE s.id=%s""",
                 (student_id,), fetchone=True)
    if not info:
        flash("Студент не найден.", "error")
        return redirect(url_for("admin_documents"))
    courses = query("""
        SELECT co.title, e.status, g.grade FROM enrollments e
        JOIN courses co ON e.course_id=co.id
        LEFT JOIN grades g ON g.enrollment_id=e.id
        WHERE e.student_id=%s ORDER BY co.title
    """, (student_id,), fetchall=True)
    _, gen_study, _ = _import_docs()
    if not gen_study:
        flash("Установите python-docx: pip install python-docx", "error")
        return redirect(url_for("admin_documents"))
    doc_bytes = gen_study(info["full_name"], info["group_name"] or "",
                          info["enrollment_year"] or datetime.now().year,
                          [dict(c) for c in courses])
    return send_file(io.BytesIO(doc_bytes), as_attachment=True,
                     download_name=f"Справка_{info['full_name'].replace(' ','_')}.docx",
                     mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document")


# ─── Run ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    app.run(debug=True, port=5000)
